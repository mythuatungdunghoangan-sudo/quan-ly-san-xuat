import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

_NHAN_C115_COLS  = ["STT", "Khách hàng", "Ngày đặt", "Ngày giao",
                    "Tên sản phẩm", "Rộng", "Cao", "Số lượng nhãn", "Ghi chú"]
_HOP_BASE_COLS   = ["STT", "Khách hàng", "Ngày đặt", "Ngày giao",
                    "Tên sản phẩm", "Số lượng hộp", "Quy cách (SP/hộp)",
                    "Kích thước hộp (DxRxC)", "Chất liệu hộp", "In ấn", "Ghi chú"]
_THUNG_BASE_COLS = ["STT", "Khách hàng", "Ngày đặt", "Ngày giao",
                    "Tên sản phẩm", "Số lượng thùng", "Số hộp/thùng",
                    "Kích thước thùng (DxRxC)", "Trọng lượng (kg)", "Ghi chú"]

TEMPLATE_COLUMNS = {
    "Nhãn C115":    _NHAN_C115_COLS,
    "Nhãn Decan":   _NHAN_C115_COLS,
    "Hộp":          _HOP_BASE_COLS,
    "Thùng carton": _THUNG_BASE_COLS,
    "Túi màng": [
        "STT", "Khách hàng", "Ngày đặt", "Ngày giao",
        "Tên sản phẩm", "Số lượng", "Đơn vị",
        "Kích thước túi (RxC)", "Chất liệu màng", "Độ dày (micron)", "Ghi chú",
    ],
    "Tổng hợp": [
        "STT", "Khách hàng", "Ngày đặt", "Ngày giao",
        "Loại", "Tên sản phẩm", "Số lượng", "Đơn vị", "Ghi chú",
    ],
}

SHEET_COLORS = {
    "Nhãn C115":    "2E75B6",  # xanh dương đậm
    "Nhãn Decan":   "9DC3E6",  # xanh dương nhạt
    "Hộp":          "70AD47",  # xanh lá
    "Thùng carton": "ED7D31",  # cam
    "Túi màng":     "FF0066",  # đỏ hồng
    "Tổng hợp":     "7030A0",  # tím
}

_COLUMN_WIDTHS = {
    "STT": 6,
    "Khách hàng": 26,
    "Ngày đặt": 13,
    "Ngày giao": 13,
    "Tên sản phẩm": 30,
    "Số lượng nhãn": 16,
    "Số lượng hộp": 16,
    "Số lượng thùng": 16,
    "Số lượng": 13,
    "Kích thước túi (RxC)": 22,
    "Chất liệu màng": 18,
    "Độ dày (micron)": 16,
    "Rộng": 10,
    "Cao": 10,
    "Kích thước hộp (DxRxC)": 24,
    "Kích thước thùng (DxRxC)": 24,
    "Chất liệu hộp": 16,
    "Quy cách (SP/hộp)": 18,
    "Số hộp/thùng": 14,
    "Trọng lượng (kg)": 16,
    "In ấn": 16,
    "Đơn vị": 10,
    "Loại": 10,
    "Ghi chú": 32,
}

_THIN = Side(style="thin", color="D0D0D0")
_DATA_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_style(color_hex):
    font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )
    return font, fill, align, border


def create_template(output_path: str | Path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in TEMPLATE_COLUMNS.items():
        ws = wb.create_sheet(sheet_name)
        color = SHEET_COLORS.get(sheet_name, "4472C4")
        font, fill, align, border = _header_style(color)

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTHS.get(col_name, 15)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 32
        ws.sheet_properties.tabColor = color

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
