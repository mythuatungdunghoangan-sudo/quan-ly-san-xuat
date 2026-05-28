import io
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from pathlib import Path

from .template_creator import TEMPLATE_COLUMNS, SHEET_COLORS

_THIN = Side(style="thin", color="D0D0D0")
_DATA_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_DATA_ALIGN = Alignment(vertical="center", wrap_text=False)
_DATA_FONT = Font(name="Calibri", size=10)
_ALT_FILL = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")


def _get_next_data_row(ws) -> int:
    """Return the first empty row after the header (row 1)."""
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return r + 1
    return 2


def _header_map(ws) -> dict[str, int]:
    """Return {header_text: col_index} for row 1."""
    mapping = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            mapping[str(val).strip()] = col
    return mapping


_NAN_VALUES = {"nan", "none", "null", "-", "n/a", ""}


def _clean(v) -> str:
    """Chuyển giá trị sang string sạch, loại bỏ nan/None."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in _NAN_VALUES else s


def _resolve_value(record: dict, target_header: str, order_info: dict) -> str:
    """Find best value for a template column from extracted record + order_info."""
    # Direct match
    for k, v in record.items():
        if str(k).strip() == target_header:
            return _clean(v)

    # Fallback: order_info fields
    _ORDER_FIELD_MAP = {
        "Mã đơn hàng": "order_id",
        "Khách hàng":  "customer",
        "Ngày đặt":    "order_date",
        "Ngày giao":   "delivery_date",
    }
    if target_header in _ORDER_FIELD_MAP:
        val = order_info.get(_ORDER_FIELD_MAP[target_header], "")
        if val:
            return _clean(val)

    # Fuzzy: partial key match
    target_low = target_header.lower()
    for k, v in record.items():
        k_low = str(k).lower()
        if k_low in target_low or target_low in k_low:
            return _clean(v)

    return ""


def append_rows(ws, records: list[dict], order_info: dict):
    hmap = _header_map(ws)
    if not hmap:
        return

    start = _get_next_data_row(ws)

    for i, record in enumerate(records):
        row = start + i
        is_alt = (row % 2 == 0)

        for header, col in hmap.items():
            cell = ws.cell(row=row, column=col)

            if header == "STT":
                cell.value = row - 1
            else:
                cell.value = _resolve_value(record, header, order_info)

            cell.font = _DATA_FONT
            cell.border = _DATA_BORDER
            cell.alignment = _DATA_ALIGN
            if is_alt:
                cell.fill = _ALT_FILL

        ws.row_dimensions[row].height = 18


def export_to_bytes(template_path: str | Path, extracted_list: list[dict]) -> bytes:
    """
    Write all extracted items to the template and return the result as bytes.
    extracted_list items: {sheet, data: list[dict]|DataFrame, order_info}
    """
    wb = openpyxl.load_workbook(template_path)

    for item in extracted_list:
        sheet_name = item["sheet"]
        order_info = item.get("order_info") or {}

        # Normalise data to list[dict]
        raw_data = item.get("data", [])
        if hasattr(raw_data, "to_dict"):
            records = raw_data.to_dict("records")
        else:
            records = list(raw_data) if raw_data else []

        if not records:
            continue

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            color = SHEET_COLORS.get(sheet_name, "4472C4")
            cols = TEMPLATE_COLUMNS.get(sheet_name, list(records[0].keys()))
            from openpyxl.styles import Font as F, PatternFill as P, Alignment as A
            header_font = F(bold=True, color="FFFFFF", size=11, name="Calibri")
            header_fill = P(start_color=color, end_color=color, fill_type="solid")
            header_align = A(horizontal="center", vertical="center")
            for c, h in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"
            ws.sheet_properties.tabColor = color

        append_rows(ws, records, order_info)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
